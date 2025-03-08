#!/usr/bin/env python3
import argparse
import asyncio
import sys
import json
import logging
import re
import pytz
from datetime import datetime, timedelta
from viam.app.viam_client import ViamClient, DataClient
from viam.rpc.dial import DialOptions, Credentials
from isodate import parse_duration, parse_datetime
from openpyxl import Workbook, load_workbook

LOGGER = logging.getLogger(__name__)
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)-8s %(message)s', datefmt='%Y-%m-%dT%H:%M:%S'))
LOGGER.addHandler(handler)


class DefaultSettings:
    def __init__(self):
        pass


class Settings(DefaultSettings):
    Command: str = None
    ApiKeyId: str = None
    ApiKey: str = None
    OrgId: str = None
    ComponentType: str = None
    LocationId: str = None
    MachineId: str = None
    ResourceName: str = None
    PartId: str = None
    Pipeline: list = None
    Start: datetime = None
    End: datetime = None
    Duration: timedelta = None
    BucketPeriod: timedelta = None
    BucketMethod: str = None
    Timezone: pytz.tzinfo = None
    OutputFile: str = None
    InputFile: str = None
    Tab: str = None
    ExcludeKeysRegex: re = None
    IncludeKeysRegex: re = None

    def __init__(self,
                 command,
                 apiKeyId,
                 apiKey,
                 orgId,
                 componentType=None,
                 locationId=None,
                 machineId=None,
                 resourceName=None,
                 partId=None,
                 pipeline=None,
                 start=None,
                 end=None,
                 duration=None,
                 output=None,
                 input=None,
                 tab=None,
                 timezone=None,
                 includeKeys=None,
                 excludeKeys=None,
                 bucketPeriod=None,
                 bucketMethod=None,
                 verbose=logging.WARN):
        super().__init__()
        self.Command = command
        self.ApiKeyId = apiKeyId
        self.ApiKey = apiKey
        self.OrgId = orgId
        self.ComponentType = componentType
        self.LocationId = locationId
        self.MachineId = machineId
        self.ResourceName = resourceName
        self.PartId = partId
        if pipeline:
            p = json.loads(pipeline)
            if not isinstance(p, list):
                raise ValueError("--pipeline must be a valid JSON array")
            self.Pipeline = p
        self.Pipeline = pipeline
        self.Start = start
        self.End = end
        self.OutputFile = output
        self.InputFile = input
        self.Tab = tab
        self.Duration = duration

        self.BucketPeriod = bucketPeriod
        if self.BucketPeriod is None and bucketMethod is not None:
            raise ValueError("--bucketMethod requires --bucketPeriod")
        self.BucketMethod = bucketMethod

        self.Timezone = timezone

        if includeKeys is not None:
            self.IncludeKeysRegex = re.compile(includeKeys)
        if excludeKeys is not None:
            self.ExcludeKeysRegex = re.compile(excludeKeys)

        if verbose >= 1:
            log_level = logging.INFO
        if verbose >= 2:
            log_level = logging.DEBUG

        LOGGER.setLevel(log_level)

        self.Timezone = parse_tzinfo(timezone)

    def build_filter(self):
        match_predicate = {}
        if self.OrgId is not None:
            match_predicate["organization_id"] = self.OrgId
        if self.LocationId is not None:
            match_predicate["location_id"] = self.LocationId
        if self.MachineId is not None:
            match_predicate["robot_id"] = self.MachineId
        if self.PartId is not None:
            match_predicate["part_id"] = self.PartId
        if self.ResourceName is not None:
            match_predicate["component_name"] = self.ResourceName
        match_predicate["time_received"] = {
            "$gte": self.Start,
            "$lt": self.End
        }
        filter = [
            {
                "$match": match_predicate
            },
            {
                "$sort": {
                    "time_received": 1
                }
            }
        ]
        LOGGER.debug(f"Built filter: {filter}")
        self.Pipeline = filter

    def add_bucket_pipeline(self):
        base_pipeline = self.Pipeline

        seconds = self.BucketPeriod.total_seconds()
        if seconds % 86400 == 0:
            unit = "day"
            amount = int(seconds / 86400)
        elif seconds % 3600 == 0:
            unit = "hour"
            amount = int(seconds / 3600)
        elif seconds % 60 == 0:
            unit = "minute"
            amount = int(seconds / 60)
        else:
            unit = "second"
            amount = int(seconds)

        bucket_stages = [
            {
                "$group": {
                    "_id": {
                        "$dateTrunc": {
                            "date": "$time_received",
                            "unit": unit,
                            "binSize": amount
                        },
                        "location_id": "$location_id",
                        "robot_id": "$robot_id",
                        "part_id": "$part_id",
                        "component_name": "$component_name",
                        "organization_id": "$organization_id",
                        "method_name": "$method_name"
                    },
                    "readings_array": {"$push": {"$objectToArray": "$data.readings"}}
                }
            },
            {"$unwind": "$readings_array"},
            {"$unwind": "$readings_array"},
            # # Group again to calculate max for each key
            # {
            #     "$group": {
            #         "_id": {
            #             "time_bucket": "$_id.time_bucket",
            #             "location_id": "$_id.location_id",
            #             "robot_id": "$_id.robot_id",
            #             "part_id": "$_id.part_id",
            #             "component_name": "$_id.component_name",
            #             "organization_id": "$_id.organization_id",
            #             "method_name": "$_id.method_name",
            #             "key": "$readings_array.k"
            #         },
            #         "time_start": {"$first": "$time_start"},
            #         "time_end": {"$first": "$time_end"},
            #         "count": {"$first": "$count"},
            #         # Apply the selected aggregation method
            #         "value": self._get_aggregation_operator("$readings_array.v")
            #     }
            # },
            # # Group to reconstruct the readings object
            # {
            #     "$group": {
            #         "_id": {
            #             "time_bucket": "$_id.time_bucket",
            #             "location_id": "$_id.location_id",
            #             "robot_id": "$_id.robot_id",
            #             "part_id": "$_id.part_id",
            #             "component_name": "$_id.component_name",
            #             "organization_id": "$_id.organization_id",
            #             "method_name": "$_id.method_name"
            #         },
            #         "time_start": {"$first": "$time_start"},
            #         "time_end": {"$first": "$time_end"},
            #         "count": {"$first": "$count"},
            #         "readings": {"$push": {"k": "$_id.key", "v": "$value"}}
            #     }
            # },
            # # Convert the array back to an object
            # {
            #     "$project": {
            #         "time_received": "$_id.time_bucket",  # Use the truncated date as time_received
            #         "part_id": "$_id.part_id",
            #         "organization_id": "$_id.organization_id",
            #         "method_name": "$_id.method_name",
            #         "location_id": "$_id.location_id",
            #         "robot_id": "$_id.robot_id",
            #         "component_name": "$_id.component_name",
            #         "data": {"readings": {"$arrayToObject": "$readings"}}
            #     }
            # },
            # # Sort by time_received
            # {"$sort": {"time_received": 1}}
        ]

        if len(base_pipeline) > 0:
            if base_pipeline[-1].get("$sort", None) is not None:
                base_pipeline.pop()

        base_pipeline.extend(bucket_stages)
        self.Pipeline = base_pipeline
        print(f"Bucket Pipeline: {self.Pipeline}")

    def _get_aggregation_operator(self, field):
        """Return the appropriate aggregation operator based on bucket method."""
        method = self.BucketMethod.lower()
        if method == "max":
            return {"$max": field}
        elif method == "min":
            return {"$min": field}
        elif method == "avg":
            return {"$avg": field}
        elif method == "first":
            return {"$first": field}
        elif method == "last":
            return {"$last": field}
        elif method == "sum":
            return {"$sum": field}
        else:
            raise ValueError(f"Unsupported bucket method: {method}")

    def validate(self):
        # Check if both pipeline and filter options are specified
        if self.Pipeline is not None and (self.ComponentType is not None or self.LocationId is not None or self.MachineId is not None or self.ResourceName is not None or self.PartId is not None):
            raise ValueError(
                "Cannot specify both --pipeline and basic filter options (--componentType, --locationId, --machineId, --resourceName, --partId)")

        # if all 3 time parameters are specified, raise an error
        if self.Start is not None and self.End is not None and self.Duration is not None:
            raise ValueError("Cannot specify --start, --end, and --duration")
        # if none are specified, set defaults
        if self.Start is None and self.End is None and self.Duration is None:
            LOGGER.debug(
                "--start, --end, and --duration not specified, defaulting to 1 day ago and now")
            self.End = datetime.now()
            self.Start = self.End - timedelta(days=1)
        # if only start and duration are specified, calculate end
        if self.Start is not None and self.Duration is not None:
            LOGGER.debug("--start and --duration specified, calculating end")
            self.End = self.Start + self.Duration
        if self.Start is not None and self.End is None and self.Duration is None:
            LOGGER.debug(
                "--end and --duration not specified, defaulting to now")
            self.End = datetime.now()
        # if only end and duration are specified, calculate start
        if self.End is not None and self.Duration is not None:
            LOGGER.debug("--end and --duration specified, calculating start")
            self.Start = self.End - self.Duration
        if self.Start is None and self.End is None and self.Duration is not None:
            LOGGER.debug("--start and --end not specified, defaulting to now")
            self.Start = datetime.now() - self.Duration
            self.End = datetime.now()

        if self.Command == 'excel':
            if self.OutputFile is None and self.InputFile is None:
                raise ValueError(
                    "--output or --input is required for Excel export")
            if self.OutputFile is not None and self.InputFile is not None:
                raise ValueError(
                    "--output and --input cannot be specified together")
            if self.InputFile is not None and self.Tab is None:
                LOGGER.warning(
                    "--input specified without --tab, will default to active tab, data loss may occurr")

        # Build the filter if the pipeline is not specified
        if self.Pipeline is None:
            LOGGER.debug("Building filter")
            self.build_filter()


class DataExporter:
    def __init__(self):
        pass

    async def ExportAsync(self):
        raise RuntimeError("ExportAsync is implemented in derived classes")

    @classmethod
    def register_with_cli(cls, subcommand: argparse.ArgumentParser):
        parser = subcommand.add_parser(
            'excel', description='Export data to Excel')
        auth_group = parser.add_argument_group('Authentication')
        auth_group.add_argument(
            '--apiKeyId', type=str, required=True, help='Viam API Key ID (required)')
        auth_group.add_argument('--apiKey', type=str,
                                required=True, help='Viam API Key (required)')

        required_filter_group = parser.add_argument_group('Required Filters')
        required_filter_group.add_argument(
            '--orgId', type=str, required=True, help='Organization ID (required)')

        standard_filter_group = parser.add_argument_group('Standard Filters')
        standard_filter_group.add_argument(
            '--componentType', type=str, help='Component Type')
        standard_filter_group.add_argument(
            '--locationId', type=str, help='Location ID')
        standard_filter_group.add_argument(
            '--machineId', type=str, help='Machine ID')
        standard_filter_group.add_argument(
            '--resourceName', type=str, help='Resource Name')
        standard_filter_group.add_argument(
            '--partId', type=str, help='Part ID')

        advanced_filter_group = parser.add_argument_group('Advanced Filters')
        advanced_filter_group.add_argument(
            '--pipeline', type=str, help='A MongoDB Pipeline to customize the data')
        advanced_filter_group.add_argument('--includeKeys', type=str,
                                           help='Include only keys matching the specified regex')
        advanced_filter_group.add_argument('--excludeKeys', type=str,
                                           help='Exclude keys matching the specified regex')

        date_group = parser.add_argument_group('Date Range')
        date_group.add_argument(
            '--start', type=parse_datetime, help='Start Date, defaults to now')
        date_group.add_argument(
            '--end', type=parse_datetime, help='End Date, defaults to 1 day ago')
        date_group.add_argument('--duration', type=parse_duration,
                                help='An ISO8601 Duration to apply to the start date')
        date_group.add_argument('--bucketPeriod', type=parse_duration,
                                help='Bucket data by time (ISO8601 Duration)')
        date_group.add_argument('--bucketMethod', type=str,
                                help='Bucketing Method for the data (min, max, avg, first, last)')

        file_group = parser.add_argument_group('File Settings')
        file_exclusive_group = file_group.add_mutually_exclusive_group(
            required=True)
        file_exclusive_group.add_argument(
            '--output', type=str, help='Output file path')
        file_exclusive_group.add_argument(
            '--input', type=str, help='Input file path')
        file_group.add_argument(
            '--tab', type=str, help='Tab name in the Excel file')

        miscellaneous_group = parser.add_argument_group('Miscellaneous')
        miscellaneous_group.add_argument(
            '--timezone', type=str, default='UTC', help='Timezone to convert timestamps to')


class Excel(DataExporter):
    client: DataClient
    settings: Settings

    def __init__(self, client: DataClient, settings: Settings):
        self.client = client
        self.settings = settings

    @classmethod
    def register_with_cli(cls, parser: argparse.ArgumentParser):
        super().register_with_cli(parser)

    async def ExportAsync(self):
        sheet_name = self.settings.Tab
        if sheet_name is None:
            sheet_name = "Sheet1"

        # Create a new workbook and select the active worksheet
        save_file = self.settings.OutputFile

        if self.settings.InputFile:
            save_file = self.settings.InputFile
            wb = load_workbook(self.settings.InputFile)
        else:
            wb = Workbook()
        
        if sheet_name in wb.sheetnames:
            wb.worksheets.remove(wb[sheet_name])

        ws = wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        wb.active = ws

        # Save the workbook to the specified output file
        wb.save(save_file)
        # we need to paginate the data now...
        skip = 0
        limit = 1000
        data = []
        while True:
            LOGGER.info(f"Retrieving data from {skip} to {skip + limit}")
            pipeline = self.settings.Pipeline.copy()
            pipeline.append({"$skip": skip})
            pipeline.append({"$limit": limit})
            LOGGER.debug(f"Executing pipeline: {pipeline}")
            batch = await self.client.tabular_data_by_mql(organization_id=self.settings.OrgId, query=pipeline)
            batchLen = len(batch)
            if batchLen == 0:  # No more data
                break
            data.extend(batch)
            if batchLen < limit:  # We less than the limit, so we are done
                break
            skip += limit
            LOGGER.info(f"Retrieved {limit} records")

        # Assuming data is a list of dictionaries, write headers
        if data:
            if self.settings.BucketPeriod is not None:
                data = bucket_data(self.settings, data)
            data_keys = sorted(data[0]["data"]["readings"].keys())
            headers = ["time_received"] + data_keys
            ws.append(headers)

            # Write data rows
            for row in data:
                time_received = row["time_received"]  # type: datetime
                utc_time_received = time_received.replace(tzinfo=pytz.utc)
                local_time_received = utc_time_received.astimezone(
                    tz=self.settings.Timezone)
                try:
                    row_values = [local_time_received.replace(
                        tzinfo=None)] + [row["data"]["readings"][key] for key in data_keys]
                    ws.append(row_values)
                except Exception as e:
                    LOGGER.warning(f"Erroring exporting row error: {e}")
                    LOGGER.warning(f"Row: {row}")

            wb.save(save_file)


def floor_timestamp(ts, bucket_td):
    # Use a fixed reference point (Unix epoch)
    epoch = datetime(1970, 1, 1)
    # Calculate the number of complete buckets since the epoch
    bucket_count = (ts - epoch) // bucket_td
    return epoch + bucket_count * bucket_td


def bucket_data(settings: Settings, data: list) -> list:
    bucketed_data = {}

    for row in data:
        time_received = row["time_received"]
        bucket = floor_timestamp(time_received, settings.BucketPeriod)
        if bucket not in bucketed_data:
            bucketed_data[bucket] = {}
        for key, value in row["data"]["readings"].items():
            if settings.ExcludeKeysRegex is not None and settings.ExcludeKeysRegex.match(key):
                continue
            if settings.IncludeKeysRegex is not None and not settings.IncludeKeysRegex.match(key):
                continue
            if key not in bucketed_data[bucket]:
                bucketed_data[bucket][key] = []
            bucketed_data[bucket][key].append(value)
    LOGGER.debug(f"Buckets: {bucketed_data.keys()}")
    aggregated_data = []
    for bucket, readings in bucketed_data.items():
        aggregated_reading = {}
        for key, values in readings.items():
            if settings.BucketMethod == "max":
                aggregated_reading[key] = max(values)
            elif settings.BucketMethod == "min":
                aggregated_reading[key] = min(values)
            elif settings.BucketMethod == "avg":
                aggregated_reading[key] = sum(values) / len(values)
            elif settings.BucketMethod == "first":
                aggregated_reading[key] = values[0]
            elif settings.BucketMethod == "last":
                aggregated_reading[key] = values[-1]
            else:
                raise ValueError(
                    f"Unsupported bucket method: {settings.BucketMethod}")
        aggregated_data.append({
            "time_received": bucket,
            "data": {"readings": aggregated_reading}
        })
    return aggregated_data


class CSV(DataExporter):
    def __init__(self):
        pass

    @classmethod
    def register_with_cli(cls, parser: argparse.ArgumentParser):
        super().register_with_cli(parser)


def parse_tzinfo(tz: str) -> pytz.tzinfo:
    return pytz.timezone(tz)


async def connect(apiKeyId: str, apiKey: str) -> ViamClient:
    dial_options = DialOptions(
        credentials=Credentials(
            type="api-key",
            payload=apiKey,
        ),
        auth_entity=apiKeyId
    )
    return await ViamClient.create_from_dial_options(dial_options)


async def main():
    parser = argparse.ArgumentParser(description='Viam Data Export')
    parser.add_argument(
        '-v', '--verbose', action='count', default=0,
        help='Increase verbosity level (e.g., -v for INFO, -vv for DEBUG)'
    )

    sub_parsers = parser.add_subparsers(required=True, dest='command')
    Excel.register_with_cli(sub_parsers)

    args = parser.parse_args()

    settings = Settings(**vars(args))

    try:
        settings.validate()
    except ValueError as e:
        parser.error(str(e))

    # Create Viam AppClient
    viam_client = await connect(settings.ApiKeyId, settings.ApiKey)
    try:
        data_client = viam_client.data_client

        if args.command == 'excel':
            exporter = Excel(data_client, settings)
            await exporter.ExportAsync()
        elif args.command == 'csv':
            exporter = CSV(data_client, settings)
            await exporter.ExportAsync()
        else:
            raise ValueError("Invalid command")
    finally:
        viam_client.close()

if __name__ == '__main__':
    asyncio.run(main())
